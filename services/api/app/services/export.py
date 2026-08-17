"""Exportación de resultados de llamadas a Excel.

Dos hojas, porque son dos preguntas distintas:

  Llamadas    una fila por llamada, una columna por pregunta. Es la vista para
              leer y filtrar en Excel.
  Respuestas  una fila por respuesta, con la transcripción textual. Formato
              largo: es la que sirve para tablas dinámicas y para revisar qué
              dijo el cliente con sus palabras.

El archivo se arma en memoria y se devuelve en la misma request. Para el volumen
de una encuesta de posventa (miles de filas al año) alcanza de sobra; el tope
de MAX_FILAS existe para que un filtro demasiado amplio no se coma la RAM del
contenedor en silencio.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import CallAttempt, QuestionType
from app.services.scoring import SATISFACTORY_MIN, is_satisfactory, to_scale_10

log = logging.getLogger(__name__)

# Tope duro. Si se alcanza, el archivo lo dice en la primera fila: un export
# truncado en silencio se lee como si fuera el total y lleva a conclusiones
# equivocadas.
MAX_FILAS = 20_000

ENCABEZADO_FONDO = PatternFill("solid", fgColor="1F2933")
ENCABEZADO_FUENTE = Font(bold=True, color="FFFFFF", size=11)

RESULTADO_ES = {
    "completed": "Completada",
    "partial": "Parcial",
    "no_answer": "No atendió",
    "busy": "Ocupado",
    "rejected": "Rechazada",
    "failed": "Error técnico",
}


def _local(momento: datetime | None, tz: ZoneInfo) -> datetime | None:
    """Excel no maneja zonas horarias: se convierte a hora local y se saca el tz."""
    if momento is None:
        return None
    if momento.tzinfo is None:
        momento = momento.replace(tzinfo=timezone.utc)
    return momento.astimezone(tz).replace(tzinfo=None)


def _valor_legible(answer) -> str | float | None:  # noqa: ANN001
    """El valor de una respuesta como lo espera ver una persona en la planilla."""
    if answer.value_numeric is None:
        return answer.transcript or None

    qtype = answer.question.qtype
    if qtype is QuestionType.YES_NO:
        return "Sí" if answer.value_numeric >= 0.5 else "No"
    if qtype is QuestionType.OPEN:
        return answer.transcript or None
    # Las escalas van como número para que Excel pueda promediarlas
    return answer.value_numeric


def _formatear(hoja, encabezados: list[str], columna: str, formato: str) -> None:
    """Aplica un formato de número a la columna que se llame `columna`."""
    try:
        indice = encabezados.index(columna) + 1
    except ValueError:
        log.warning("No existe la columna %r; no se aplicó el formato", columna)
        return

    for fila in hoja.iter_rows(min_row=2, min_col=indice, max_col=indice):
        fila[0].number_format = formato


def _ajustar(hoja, anchos: dict[int, int]) -> None:
    """Ancho de columnas, panel fijo y autofiltro: lo que hace usable la hoja."""
    for indice, ancho in anchos.items():
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    for celda in hoja[1]:
        celda.fill = ENCABEZADO_FONDO
        celda.font = ENCABEZADO_FUENTE
        celda.alignment = Alignment(vertical="center", wrap_text=True)

    hoja.row_dimensions[1].height = 30
    hoja.freeze_panes = "A2"
    if hoja.max_row > 1:
        hoja.auto_filter.ref = hoja.dimensions


def _hoja_llamadas(wb: Workbook, calls: list[CallAttempt], tz: ZoneInfo) -> None:
    hoja = wb.active
    hoja.title = "Llamadas"

    # Las columnas de preguntas salen de los datos filtrados, no de un listado
    # fijo: así el export refleja exactamente las campañas que se exportaron.
    preguntas: dict[int, object] = {}
    for call in calls:
        for answer in call.answers:
            preguntas.setdefault(answer.question_id, answer.question)

    ordenadas = sorted(
        preguntas.values(), key=lambda q: (q.campaign_id, q.position)
    )

    encabezados = [
        "Fecha", "Cliente", "Teléfono", "Campaña",
        "Resultado", "Intento", "Duración (seg)",
        "Puntaje /10", "Conforme", "Sentimiento",
        "Requiere seguimiento", "Motivo", "Resumen",
        "Registro Bitrix", "Ingreso al taller",
    ]
    encabezados += [f"P{q.position}. {q.text[:60]}" for q in ordenadas]
    hoja.append(encabezados)

    for call in calls:
        target = call.target
        analisis = call.analysis
        respuestas = {a.question_id: a for a in call.answers}

        puntaje = analisis.satisfaction_score if analisis else None

        fila = [
            _local(call.started_at, tz),
            target.contact_name if target else None,
            target.phone if target else None,
            target.campaign.name if target and target.campaign else None,
            RESULTADO_ES.get(call.outcome.value, call.outcome.value) if call.outcome else "En curso",
            call.attempt_number,
            call.duration_seconds,
            puntaje,
            # Columna explícita en vez de dejar que cada quien recuerde el umbral
            ("Sí" if is_satisfactory(puntaje) else "No") if puntaje is not None else None,
            analisis.sentiment if analisis else None,
            ("Sí" if analisis.requires_followup else "No") if analisis else None,
            analisis.followup_reason if analisis else None,
            analisis.summary if analisis else None,
            f"{target.bitrix_entity_type_id}#{target.bitrix_entity_id}" if target else None,
            _local(target.trigger_at, tz) if target else None,
        ]
        fila += [
            _valor_legible(respuestas[q.id]) if q.id in respuestas else None
            for q in ordenadas
        ]
        hoja.append(fila)

    # Los formatos se ubican por nombre de encabezado, no por número de columna:
    # con índices fijos, reordenar una columna rompe las fechas sin que nada
    # falle — el archivo sale con números crudos donde iban fechas.
    _formatear(hoja, encabezados, "Fecha", "dd/mm/yyyy hh:mm")
    _formatear(hoja, encabezados, "Ingreso al taller", "dd/mm/yyyy hh:mm")
    _formatear(hoja, encabezados, "Puntaje /10", "0.0")

    _ajustar(hoja, {
        1: 17, 2: 26, 3: 16, 4: 24, 5: 14, 6: 8, 7: 14,
        8: 11, 9: 10, 10: 13, 11: 12, 12: 40, 13: 46, 14: 16, 15: 17,
        **{16 + i: 22 for i in range(len(ordenadas))},
    })


def _hoja_respuestas(wb: Workbook, calls: list[CallAttempt], tz: ZoneInfo) -> None:
    hoja = wb.create_sheet("Respuestas")
    encabezados = [
        "Fecha", "Cliente", "Teléfono", "Campaña",
        "Nº", "Pregunta", "Tipo",
        "Valor", "Puntaje /10", "Conforme",
        "Lo que dijo el cliente", "Confianza ASR", "Repreguntas",
    ]
    hoja.append(encabezados)

    for call in calls:
        target = call.target
        for answer in sorted(call.answers, key=lambda a: a.question.position):
            pregunta = answer.question
            en_escala = (
                to_scale_10(answer.value_numeric, pregunta.qtype)
                if answer.value_numeric is not None
                else None
            )

            hoja.append([
                _local(call.started_at, tz),
                target.contact_name if target else None,
                target.phone if target else None,
                target.campaign.name if target and target.campaign else None,
                pregunta.position,
                pregunta.text,
                pregunta.qtype.value,
                _valor_legible(answer),
                en_escala,
                # Solo tiene sentido en las preguntas que puntúan
                ("Sí" if is_satisfactory(en_escala) else "No")
                if en_escala is not None and pregunta.counts_for_score
                else None,
                answer.transcript,
                answer.asr_confidence,
                answer.retries_used,
            ])

    _formatear(hoja, encabezados, "Fecha", "dd/mm/yyyy hh:mm")
    _formatear(hoja, encabezados, "Confianza ASR", "0%")
    _formatear(hoja, encabezados, "Puntaje /10", "0.0")

    _ajustar(hoja, {
        1: 17, 2: 26, 3: 16, 4: 22, 5: 5, 6: 52, 7: 15,
        8: 12, 9: 11, 10: 10, 11: 60, 12: 12, 13: 11,
    })


def construir_excel(
    calls: list[CallAttempt],
    tz: ZoneInfo,
    descripcion_filtros: str = "",
) -> tuple[bytes, bool]:
    """Devuelve (bytes del .xlsx, si se truncó).

    Recibe las llamadas ya filtradas y con las relaciones cargadas. Si acá se
    consultara la base se duplicaría la lógica de filtros y el export podría
    dejar de coincidir con lo que la pantalla muestra.
    """
    truncado = len(calls) > MAX_FILAS
    if truncado:
        log.warning(
            "Export truncado: %d llamadas superan el tope de %d",
            len(calls), MAX_FILAS,
        )
        calls = calls[:MAX_FILAS]

    wb = Workbook()
    _hoja_llamadas(wb, calls, tz)
    _hoja_respuestas(wb, calls, tz)

    hoja = wb.create_sheet("Acerca de", 0)
    generado = datetime.now(tz).strftime("%d/%m/%Y %H:%M")
    filas = [
        ("Encuesta de satisfacción post-taller", ""),
        ("", ""),
        ("Generado", generado),
        ("Llamadas exportadas", len(calls)),
        ("Filtros aplicados", descripcion_filtros or "ninguno (todo el histórico)"),
        ("", ""),
        ("Umbral de conformidad", f"{SATISFACTORY_MIN:.0f} de 10 o más"),
        ("", "Un puntaje por debajo del umbral marca la llamada para seguimiento."),
        ("", ""),
        ("Hoja «Llamadas»", "Una fila por llamada, una columna por pregunta."),
        ("Hoja «Respuestas»", "Una fila por respuesta, con lo que dijo el cliente."),
    ]
    if truncado:
        filas += [
            ("", ""),
            ("ATENCIÓN", f"El resultado supera {MAX_FILAS} filas y se recortó."),
            ("", "Acotá el rango de fechas para exportar el resto."),
        ]
    for fila in filas:
        hoja.append(list(fila))

    for celda in hoja["A"]:
        celda.font = Font(bold=True)
    hoja.column_dimensions["A"].width = 26
    hoja.column_dimensions["B"].width = 68
    for fila in hoja.iter_rows(min_col=2, max_col=2):
        fila[0].alignment = Alignment(wrap_text=True, vertical="top")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), truncado
